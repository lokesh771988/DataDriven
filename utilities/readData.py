import openpyxl


def getRowCount(path, sheet):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet]
    return sheet.max_row


def getColumnsCount(path, sheet):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet]
    return sheet.max_column


def getReadData(path, sheet, r, c):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet]
    return sheet.cell(row=r, column=c).value


def writeData(path, sheet, r, c, result):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet]
    sheet.cell(row=r, column=c).value = result
    workbook.save(path)
